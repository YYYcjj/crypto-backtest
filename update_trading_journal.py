"""
交易记录自动化 — 从 OKX API 拉取已平仓订单，自动生成 Excel 日记
用法: python update_trading_journal.py [月份]
      不带参数：当前月
      带参数：python update_trading_journal.py 6  (6月)
"""
import json, sys, os
from datetime import datetime, timezone, timedelta
from collections import defaultdict
import requests, hmac, base64, hashlib

# ── 配置 ──
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "results")
DATA_FILE = os.path.join(OUTPUT_DIR, "journal_trades.json")
EXCEL_FILE = os.path.join(OUTPUT_DIR, "交易记录_{month}.xlsx")
MCP_CONFIG = os.path.expanduser("~/.workbuddy/mcp.json")

os.makedirs(OUTPUT_DIR, exist_ok=True)


def get_month_range(year, month):
    """返回指定月的开始和结束时间戳（毫秒）"""
    start = int(datetime(year, month, 1, tzinfo=timezone.utc).timestamp() * 1000)
    if month == 12:
        end = int(datetime(year + 1, 1, 1, tzinfo=timezone.utc).timestamp() * 1000) - 1
    else:
        end = int(datetime(year, month + 1, 1, tzinfo=timezone.utc).timestamp() * 1000) - 1
    return start, end


def get_okx_credentials():
    """获取 OKX API 凭证：优先环境变量，fallback MCP 配置"""
    # GitHub Actions 环境变量
    if os.environ.get("OKX_API_KEY"):
        return {
            "OKX_API_KEY": os.environ["OKX_API_KEY"],
            "OKX_API_SECRET": os.environ["OKX_API_SECRET"],
            "OKX_API_PASSPHRASE": os.environ["OKX_API_PASSPHRASE"],
        }
    # 本地 MCP 配置
    if os.path.exists(MCP_CONFIG):
        with open(MCP_CONFIG) as f:
            return json.load(f)["mcpServers"]["okx-mcp"]["env"]
    raise RuntimeError("未找到 OKX API 凭证。请设置环境变量或配置 MCP。")


def okx_get(path):
    """调用 OKX API v5"""
    env = get_okx_credentials()
    now = datetime.now(timezone.utc)
    ts = now.strftime("%Y-%m-%dT%H:%M:%S.") + f"{now.microsecond // 1000:03d}Z"
    sign = base64.b64encode(
        hmac.new(env["OKX_API_SECRET"].encode(), (ts + "GET" + path).encode(), hashlib.sha256).digest()
    ).decode()
    return requests.get("https://www.okx.com" + path, headers={
        "OK-ACCESS-KEY": env["OKX_API_KEY"],
        "OK-ACCESS-SIGN": sign,
        "OK-ACCESS-TIMESTAMP": ts,
        "OK-ACCESS-PASSPHRASE": env["OKX_API_PASSPHRASE"],
    }, timeout=30).json()


def fetch_trades(start_ts, end_ts):
    """从 OKX 拉取已平仓订单并重建交易（支持分页）"""
    orders = []
    after = ""
    page = 1
    while True:
        path = f"/api/v5/trade/orders-history?instType=SWAP&state=filled&begin={start_ts}&end={end_ts}&limit=100"
        if after:
            path += f"&after={after}"
        result = okx_get(path)
        page_data = result.get("data", [])
        orders.extend(page_data)
        if len(page_data) < 100:
            break
        after = page_data[-1].get("ordId", "")
        page += 1
        if page > 20:
            break
    orders.sort(key=lambda x: int(x["cTime"]))



def enrich_with_indicators(trades):
    """为每笔交易补充入场时的 1H/4H/1D DMI方向 + SRSI"""
    try:
        sys.path.insert(0, SCRIPT_DIR)
        from data_fetcher import fetch_historical
        from indicators import calc_dmi_adx, calc_stoch_rsi
    except ImportError:
        print("⚠️ 无法导入指标模块，跳过指标计算")
        return trades

    cache = {}

    def get_candles(sym, tf, days):
        key = (sym + "-SWAP", tf)
        if key not in cache:
            c = fetch_historical(sym + "-SWAP", tf, days)
            if c:
                cache[key] = c
        return cache.get(key, [])

    def calc(candles, idx):
        if idx < 20:
            return "?", "?"
        w = candles[max(0, idx - 100): idx + 1]
        closes = [c["c"] for c in w]
        dmi, _, _ = calc_dmi_adx(w, 14)
        k, _ = calc_stoch_rsi(closes, 10, 10, 3, 3)
        return dmi, round(k, 1)

    for t in trades:
        ets = t["entry_ts"]
        for tf_name, bar, days in [("1H", "1H", 30), ("4H", "4H", 60), ("1D", "1D", 120)]:
            c = get_candles(t["coin"], bar, days)
            if not c:
                t["trend_" + tf_name.lower()] = "?"
                t["srsi_" + tf_name.lower()] = "?"
                continue
            idx = min(range(len(c)), key=lambda x: abs(int(c[x]["ts"]) - ets))
            dmi, srsi = calc(c, idx)
            t["trend_" + tf_name.lower()] = dmi
            t["srsi_" + tf_name.lower()] = srsi

    return trades


def load_existing(path):
    """加载已有交易记录"""
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return []


def merge_trades(existing, new_trades):
    """合并去重（按 entry_time + coin + entry_price 判断）"""
    existing_keys = {(t["entry_time"], t["coin"], t["entry_price"]) for t in existing}
    added = []
    for t in new_trades:
        key = (t["entry_time"], t["coin"], t["entry_price"])
        if key not in existing_keys:
            existing.append(t)
            existing_keys.add(key)
            added.append(t)
    existing.sort(key=lambda x: x["entry_ts"])
    return existing, added


def generate_excel(trades, output_path, month_label):
    """生成 Excel — 清爽配色 + 顶部可视化汇总"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month_label}月交易记录"

    # ── 配色 ──
    WHITE = "FFFFFF"
    BG_STRIPE = "f8f9fa"
    HEADER_BG = "e9ecef"
    WIN_GREEN = "198754"
    LOSE_RED = "dc3545"
    SUCCESS_BG = "d1e7dd"
    DANGER_BG = "f8d7da"
    TEXT_MAIN = "212529"
    TEXT_MUTED = "6c757d"
    BORDER_COLOR = "dee2e6"

    hdr_fill = PatternFill("solid", fgColor=HEADER_BG)
    border = Border(bottom=Side("thin", BORDER_COLOR))
    hdr_border = Border(bottom=Side("medium", "0d6efd"))
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(vertical="center")

    total_pnl = sum(t["pnl_usdt"] for t in trades)
    wins_list = [t for t in trades if t["pnl_usdt"] > 0]
    losses_list = [t for t in trades if t["pnl_usdt"] <= 0]
    win_rate = len(wins_list) / len(trades) * 100 if trades else 0
    avg_win = sum(t["pnl_usdt"] for t in wins_list) / len(wins_list) if wins_list else 0
    avg_loss = sum(t["pnl_usdt"] for t in losses_list) / len(losses_list) if losses_list else 0
    best = max(trades, key=lambda t: t["pnl_usdt"]) if trades else None
    worst = min(trades, key=lambda t: t["pnl_usdt"]) if trades else None

    # ═══════════ 顶部可视化 ═══════════
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.cell(1, 1, f"📊 {month_label}月交易记录").font = Font(bold=True, size=18, color=TEXT_MAIN)
    ws.row_dimensions[1].height = 36

    # ── 统计卡片 ──
    cards = [
        ("总交易", f"{len(trades)} 笔"),
        ("胜率", f"{win_rate:.0f}%"),
        ("净P&L", f"{total_pnl:+.1f} USDT"),
        ("最大盈利", f"+{best['pnl_usdt']:.1f}U {best['coin']}" if best else "-"),
        ("最大亏损", f"{worst['pnl_usdt']:.1f}U {worst['coin']}" if worst else "-"),
        ("盈亏比", f"{abs(avg_win/avg_loss):.1f}:1" if avg_loss else "-"),
    ]
    card_colors = [TEXT_MAIN, (WIN_GREEN if win_rate>=50 else LOSE_RED), (WIN_GREEN if total_pnl>0 else LOSE_RED), WIN_GREEN, LOSE_RED, WIN_GREEN]

    for j, ((label, value), color) in enumerate(zip(cards, card_colors)):
        col = 1 + j * 3
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        ws.cell(3, col, label).font = Font(size=9, color=TEXT_MUTED, bold=True)
        ws.cell(3, col).alignment = center
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        ws.cell(4, col, value).font = Font(size=14, color=color, bold=True)
        ws.cell(4, col).alignment = center
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 24

    # ── 品种 P&L 柱状图 ──
    coin_stats = {}
    for t in trades:
        c = t["coin"]
        if c not in coin_stats:
            coin_stats[c] = {"pnl": 0, "count": 0, "wins": 0}
        coin_stats[c]["pnl"] += t["pnl_usdt"]
        coin_stats[c]["count"] += 1
        if t["pnl_usdt"] > 0:
            coin_stats[c]["wins"] += 1

    sorted_coins = sorted(coin_stats.items(), key=lambda x: x[1]["pnl"], reverse=True)
    CD = 7  # Chart data row
    ws.cell(CD, 1, "品种").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(CD, 2, "P&L(USDT)").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(CD, 3, "笔数").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(CD, 4, "胜率").font = Font(size=9, color=TEXT_MUTED)

    for i, (coin, stats) in enumerate(sorted_coins):
        r = CD + 1 + i
        ws.cell(r, 1, coin).font = Font(size=10, color=TEXT_MAIN)
        ws.cell(r, 2, round(stats["pnl"], 1)).font = Font(size=10)
        ws.cell(r, 3, stats["count"]).font = Font(size=10, color=TEXT_MUTED)
        wr = stats["wins"] / stats["count"] * 100
        ws.cell(r, 4, f"{wr:.0f}%").font = Font(size=10, color=WIN_GREEN if wr>=50 else LOSE_RED)

    CEND = CD + len(sorted_coins)

    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.title = None
    chart.y_axis.title = "P&L (USDT)"
    chart.legend = None
    chart.width = 18
    chart.height = 10

    dref = Reference(ws, min_col=2, min_row=CD, max_col=2, max_row=CEND)
    cref = Reference(ws, min_col=1, min_row=CD+1, max_row=CEND)
    chart.add_data(dref, titles_from_data=True)
    chart.set_categories(cref)
    chart.shape = 4

    from openpyxl.chart.series import DataPoint
    for i in range(len(sorted_coins)):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = WIN_GREEN if sorted_coins[i][1]["pnl"] > 0 else LOSE_RED
        chart.series[0].data_points.append(pt)

    ws.add_chart(chart, f"F{CD}")

    # Hide chart data
    for r in range(CD, CEND + 1):
        for c in range(1, 5):
            ws.cell(r, c).value = None

    # ═══════════ 明细表 ═══════════
    TS = CEND + 2  # Table start

    headers1 = ["交易日期","Coin","入场判断","","","","","","多/空","入场截图","更多-1","更多-2","判断","离场截图","盈亏比","交易得失分析","整体截图","",""]
    headers2 = ["","","趋势方向","","","SRSI","","","","","","","","","","","","",""]
    headers3 = ["","","1H","4H","1D","1H","4H","1D","","","","","","","","","","",""]

    for hr, hd in enumerate([headers1, headers2, headers3]):
        row = TS + hr
        for col, val in enumerate(hd, 1):
            c = ws.cell(row, col, val)
            c.fill = hdr_fill
            c.font = Font(bold=True, color=TEXT_MUTED, size=9)
            c.alignment = center
            c.border = hdr_border
        ws.row_dimensions[row].height = 18

    DS = TS + 3  # Data start
    for i, t in enumerate(trades):
        row = DS + i
        pnl = t["pnl_usdt"]

        ws.cell(row, 1, t["entry_time"][:16]).font = Font(size=10, color=TEXT_MAIN)
        ws.cell(row, 2, t["coin"]).font = Font(bold=True, size=10, color=TEXT_MAIN)

        for j, k in enumerate(["trend_1h", "trend_4h", "trend_1d"]):
            v = t.get(k, "?")
            ws.cell(row, 3+j, v).font = Font(color=WIN_GREEN if v=="多" else (LOSE_RED if v=="空" else TEXT_MUTED), size=10, bold=True)

        for j, k in enumerate(["srsi_1h", "srsi_4h", "srsi_1d"]):
            v = t.get(k, "?")
            if isinstance(v, (int, float)):
                ws.cell(row, 6+j, v).font = Font(color=LOSE_RED if v>80 else (WIN_GREEN if v<20 else "f59f00"), size=10)
            else:
                ws.cell(row, 6+j, v).font = Font(color=TEXT_MUTED, size=10)

        ws.cell(row, 9, t["direction"]).font = Font(color=WIN_GREEN if "多" in t["direction"] else LOSE_RED, bold=True, size=10)

        pct = t.get("pnl_pct", 0)
        ws.cell(row, 16, f"{pct:+.1f}%").font = Font(color=WIN_GREEN if pnl>0 else LOSE_RED, bold=True, size=10)

        note = "大盈利" if pnl>10 else ("盈利" if pnl>3 else ("小盈" if pnl>0 else ("小损" if pnl>-4 else "止损")))
        ws.cell(row, 17, note).font = Font(size=9, color=WIN_GREEN if pnl>0 else LOSE_RED)

        row_fill = PatternFill("solid", fgColor=BG_STRIPE) if i%2==0 else PatternFill("solid", fgColor=WHITE)
        for col in range(1, 20):
            ws.cell(row, col).fill = row_fill
            ws.cell(row, col).border = border
            ws.cell(row, col).alignment = left

        ws.cell(row, 16).fill = PatternFill("solid", fgColor=SUCCESS_BG if pnl>0 else DANGER_BG)

        # Add MAE/MFE columns (cols 18, 19)
        mae = t.get("mae_pct", 0)
        mfe = t.get("mfe_pct", 0)
        ws.cell(row, 18, f"{mae:+.1f}%").font = Font(size=9, color=LOSE_RED)
        ws.cell(row, 19, f"{mfe:+.1f}%").font = Font(size=9, color=WIN_GREEN)

    # Add MAE/MFE to header
    ws.cell(TS, 18, "MAE").font = Font(bold=True, color=TEXT_MUTED, size=9)
    ws.cell(TS, 18).fill = hdr_fill
    ws.cell(TS, 18).alignment = center
    ws.cell(TS, 18).border = hdr_border
    ws.cell(TS, 19, "MFE").font = Font(bold=True, color=TEXT_MUTED, size=9)
    ws.cell(TS, 19).fill = hdr_fill
    ws.cell(TS, 19).alignment = center
    ws.cell(TS, 19).border = hdr_border

    # Column widths
    for col, w in {1:16,2:10,3:5,4:5,5:5,6:6,7:6,8:6,9:6,16:10,17:10}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = ws.cell(DS, 1)

    # ═══════════ Sheet 2: 可视化仪表盘 ═══════════
    ws2 = wb.create_sheet("可视化")
    _build_viz_sheet(ws2, trades, month_label)

    # ═══════════ Sheet 3: 走势图 ═══════════
    ws3 = wb.create_sheet("走势图")
    _build_price_paths(ws3, trades)

    wb.save(output_path)

    # Also generate web dashboard
    html_path = os.path.join(os.path.dirname(output_path), "index.html")
    _generate_html(trades, month_label, html_path)

    return output_path


def _build_viz_sheet(ws, trades, month_label):
    """第二页：多图表可视化"""
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference, ScatterChart
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter

    WHITE = "FFFFFF"
    BG = "f8f9fa"
    WIN_GREEN = "198754"
    LOSE_RED = "dc3545"
    TEXT_MAIN = "212529"
    TEXT_MUTED = "6c757d"

    ws.sheet_properties.tabColor = "0d6efd"

    ws.merge_cells("A1:G1")
    ws.cell(1, 1, f"📊 {month_label}月交易可视化").font = Font(bold=True, size=16, color=TEXT_MAIN)
    ws.row_dimensions[1].height = 30

    # ── Chart 1: 日 P&L 走势 ──
    ws.cell(3, 1, "日期").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(3, 2, "日P&L").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(3, 3, "累计P&L").font = Font(size=9, color=TEXT_MUTED)

    from collections import defaultdict
    daily = defaultdict(float)
    for t in trades:
        day = t["entry_time"][:10]
        daily[day] += t["pnl_usdt"]

    sorted_days = sorted(daily.items())
    cum = 0
    for i, (day, pnl) in enumerate(sorted_days):
        cum += pnl
        ws.cell(4 + i, 1, day[5:]).font = Font(size=9)
        ws.cell(4 + i, 2, round(pnl, 1)).font = Font(size=9)
        ws.cell(4 + i, 3, round(cum, 1)).font = Font(size=9)
    D_END = 3 + len(sorted_days)

    line_chart = LineChart()
    line_chart.title = "累计 P&L 走势"
    line_chart.y_axis.title = "USDT"
    line_chart.style = 2
    line_chart.width = 22
    line_chart.height = 12
    dref = Reference(ws, min_col=3, min_row=3, max_col=3, max_row=D_END)
    cref = Reference(ws, min_col=1, min_row=4, max_row=D_END)
    line_chart.add_data(dref, titles_from_data=True)
    line_chart.set_categories(cref)
    line_chart.series[0].graphicalProperties.solidFill = "0d6efd"
    line_chart.series[0].graphicalProperties.line.solidFill = "0d6efd"
    ws.add_chart(line_chart, "E3")

    # Hide data
    for r in range(3, D_END + 1):
        ws.cell(r, 1).value = None
        ws.cell(r, 2).value = None
        ws.cell(r, 3).value = None

    # ── Chart 2: 胜/负 饼图 ──
    PIE_ROW = D_END + 2
    wins = sum(1 for t in trades if t["pnl_usdt"] > 0)
    losses = len(trades) - wins
    ws.cell(PIE_ROW, 1, "结果").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(PIE_ROW, 2, "笔数").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(PIE_ROW + 1, 1, "盈利").font = Font(size=10)
    ws.cell(PIE_ROW + 1, 2, wins).font = Font(size=10)
    ws.cell(PIE_ROW + 2, 1, "亏损").font = Font(size=10)
    ws.cell(PIE_ROW + 2, 2, losses).font = Font(size=10)

    pie = PieChart()
    pie.title = f"胜率 {wins/len(trades)*100:.0f}%"
    pie.width = 14
    pie.height = 12
    pref = Reference(ws, min_col=2, min_row=PIE_ROW, max_col=2, max_row=PIE_ROW+2)
    cref2 = Reference(ws, min_col=1, min_row=PIE_ROW+1, max_row=PIE_ROW+2)
    pie.add_data(pref, titles_from_data=True)
    pie.set_categories(cref2)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pt0 = DataPoint(idx=0)
    pt0.graphicalProperties.solidFill = WIN_GREEN
    pt1 = DataPoint(idx=1)
    pt1.graphicalProperties.solidFill = LOSE_RED
    pie.series[0].data_points = [pt0, pt1]
    ws.add_chart(pie, f"E{PIE_ROW}")

    # Hide
    for r in range(PIE_ROW, PIE_ROW + 3):
        ws.cell(r, 1).value = None
        ws.cell(r, 2).value = None

    # ── Chart 3: 方向 vs 结果 ──
    DIR_ROW = PIE_ROW + 4
    ws.cell(DIR_ROW, 1, "方向").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(DIR_ROW, 2, "盈利").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(DIR_ROW, 3, "亏损").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(DIR_ROW, 4, "胜率").font = Font(size=9, color=TEXT_MUTED)

    for j, direction in enumerate(["做多", "做空"]):
        subset = [t for t in trades if t["direction"] == direction]
        w = sum(1 for t in subset if t["pnl_usdt"] > 0)
        l = len(subset) - w
        wr = w / len(subset) * 100 if subset else 0
        ws.cell(DIR_ROW+1+j, 1, direction).font = Font(size=10)
        ws.cell(DIR_ROW+1+j, 2, w).font = Font(size=10, color=WIN_GREEN)
        ws.cell(DIR_ROW+1+j, 3, l).font = Font(size=10, color=LOSE_RED)
        ws.cell(DIR_ROW+1+j, 4, f"{wr:.0f}%").font = Font(size=10, bold=True, color=WIN_GREEN if wr>=50 else LOSE_RED)

    bar2 = BarChart()
    bar2.type = "col"
    bar2.style = 2
    bar2.title = "做多 vs 做空"
    bar2.width = 14
    bar2.height = 12
    bar2.legend = None
    dref2 = Reference(ws, min_col=2, min_row=DIR_ROW, max_col=2, max_row=DIR_ROW+2)
    cref3 = Reference(ws, min_col=1, min_row=DIR_ROW+1, max_row=DIR_ROW+2)
    bar2.add_data(dref2, titles_from_data=True)
    bar2.set_categories(cref3)
    ws.add_chart(bar2, f"E{DIR_ROW}")

    for r in range(DIR_ROW, DIR_ROW + 3):
        for c in range(1, 5):
            ws.cell(r, c).value = None

    # ── Chart 4: 趋势对齐 vs 存活率 ──
    TREND_ROW = DIR_ROW + 4
    ws.cell(TREND_ROW, 1, "趋势对齐").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(TREND_ROW, 2, "笔数").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(TREND_ROW, 3, "胜率").font = Font(size=9, color=TEXT_MUTED)

    trend_groups = {"全多":[], "全空":[], "混合":[]}
    for t in trades:
        t1 = t.get("trend_1h", "?")
        t4 = t.get("trend_4h", "?")
        td = t.get("trend_1d", "?")
        if t1 == "多" and t4 == "多" and td == "多":
            trend_groups["全多"].append(t)
        elif t1 == "空" and t4 == "空" and td == "空":
            trend_groups["全空"].append(t)
        else:
            trend_groups["混合"].append(t)

    for j, (label, subset) in enumerate(trend_groups.items()):
        w = sum(1 for t in subset if t["pnl_usdt"] > 0)
        wr = w / len(subset) * 100 if subset else 0
        ws.cell(TREND_ROW+1+j, 1, label).font = Font(size=10)
        ws.cell(TREND_ROW+1+j, 2, len(subset)).font = Font(size=10)
        ws.cell(TREND_ROW+1+j, 3, f"{wr:.0f}%").font = Font(size=10, bold=True, color=WIN_GREEN if wr>=50 else LOSE_RED)

    bar3 = BarChart()
    bar3.type = "bar"
    bar3.style = 2
    bar3.title = "趋势对齐胜率"
    bar3.width = 14
    bar3.height = 10
    bar3.legend = None
    dref3 = Reference(ws, min_col=3, min_row=TREND_ROW, max_col=3, max_row=TREND_ROW+3)
    cref4 = Reference(ws, min_col=1, min_row=TREND_ROW+1, max_row=TREND_ROW+3)
    bar3.add_data(dref3, titles_from_data=True)
    bar3.set_categories(cref4)
    ws.add_chart(bar3, f"E{TREND_ROW}")

    for r in range(TREND_ROW, TREND_ROW + 4):
        for c in range(1, 4):
            ws.cell(r, c).value = None

    # ── Chart 5: P&L 分布 ──
    DIST_ROW = TREND_ROW + 5
    ws.cell(DIST_ROW, 1, "盈亏区间").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(DIST_ROW, 2, "笔数").font = Font(size=9, color=TEXT_MUTED)

    buckets = {"<-10":0, "-10~-5":0, "-5~0":0, "0~+5":0, "+5~+15":0, ">+15":0}
    for t in trades:
        p = t["pnl_usdt"]
        if p < -10: buckets["<-10"] += 1
        elif p < -5: buckets["-10~-5"] += 1
        elif p < 0: buckets["-5~0"] += 1
        elif p < 5: buckets["0~+5"] += 1
        elif p < 15: buckets["+5~+15"] += 1
        else: buckets[">+15"] += 1

    bucket_colors = [LOSE_RED, LOSE_RED, LOSE_RED, WIN_GREEN, WIN_GREEN, WIN_GREEN]
    for j, (label, count) in enumerate(buckets.items()):
        ws.cell(DIST_ROW+1+j, 1, label).font = Font(size=10)
        ws.cell(DIST_ROW+1+j, 2, count).font = Font(size=10, bold=True, color=TEXT_MAIN)

    bar4 = BarChart()
    bar4.type = "col"
    bar4.style = 2
    bar4.title = "P&L 分布"
    bar4.width = 14
    bar4.height = 10
    bar4.legend = None
    dref4 = Reference(ws, min_col=2, min_row=DIST_ROW, max_col=2, max_row=DIST_ROW+6)
    cref5 = Reference(ws, min_col=1, min_row=DIST_ROW+1, max_row=DIST_ROW+6)
    bar4.add_data(dref4, titles_from_data=True)
    bar4.set_categories(cref5)
    for j, bc in enumerate(bucket_colors):
        pt = DataPoint(idx=j)
        pt.graphicalProperties.solidFill = bc
        bar4.series[0].data_points.append(pt)
    ws.add_chart(bar4, f"E{DIST_ROW}")

    for r in range(DIST_ROW, DIST_ROW + 7):
        for c in range(1, 3):
            ws.cell(r, c).value = None

    # ── Add MAE vs MFE scatter ──
    MAE_ROW = DIST_ROW + 8
    ws.cell(MAE_ROW, 1, "MAE%").font = Font(size=9, color=TEXT_MUTED)
    ws.cell(MAE_ROW, 2, "MFE%").font = Font(size=9, color=TEXT_MUTED)
    for i, t in enumerate(trades):
        ws.cell(MAE_ROW+1+i, 1, t.get("mae_pct", 0)).font = Font(size=9)
        ws.cell(MAE_ROW+1+i, 2, t.get("mfe_pct", 0)).font = Font(size=9)

    scatter = ScatterChart()
    scatter.title = "MAE vs MFE (每笔交易)"
    scatter.x_axis.title = "最大亏损 (%)"
    scatter.y_axis.title = "最大盈利 (%)"
    scatter.width = 16
    scatter.height = 12
    sref = Reference(ws, min_col=1, min_row=MAE_ROW, max_col=2, max_row=MAE_ROW+len(trades))
    scatter.add_data(sref, titles_from_data=True)
    s1 = scatter.series[0]
    s1.graphicalProperties.solidFill = LOSE_RED
    s2 = scatter.series[1]
    s2.graphicalProperties.solidFill = WIN_GREEN
    ws.add_chart(scatter, f"E{MAE_ROW}")

    for r in range(MAE_ROW, MAE_ROW+len(trades)+1):
        ws.cell(r, 1).value = None; ws.cell(r, 2).value = None

    # ── Column widths ──
    for c in range(1, 5):
        ws.column_dimensions[get_column_letter(c)].width = 12

def _build_price_paths(ws, trades):
    """第三页：关键交易的持仓期间价格走势"""
    import sys, os
    from openpyxl.styles import Font
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timezone
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from data_fetcher import fetch_historical

    WS = "FFFFFF"
    WIN_GREEN = "198754"
    LOSE_RED = "dc3545"
    TEXT_MAIN = "212529"
    TEXT_MUTED = "6c757d"

    ws.sheet_properties.tabColor = "198754"
    ws.merge_cells("A1:I1")
    ws.cell(1, 1, "📈 持仓期间价格走势").font = Font(bold=True, size=16, color=TEXT_MAIN)
    ws.row_dimensions[1].height = 30

    # Pick 4 key trades: best win, worst loss, worst MAE, most recent
    best_win = max(trades, key=lambda t: t["pnl_usdt"])
    worst_loss = min(trades, key=lambda t: t["pnl_usdt"])
    worst_mae = min(trades, key=lambda t: t.get("mae_pct", 0))
    picks = [
        (best_win, f"最大盈利 {best_win['coin']} +{best_win['pnl_usdt']:.1f}U"),
        (min(trades, key=lambda t: t.get("mae_pct", 0)), f"最深回撤 {worst_mae['coin']} MAE={worst_mae.get('mae_pct',0):.1f}%"),
        (worst_loss, f"最大亏损 {worst_loss['coin']} {worst_loss['pnl_usdt']:.1f}U"),
        (trades[-1], f"最近 {trades[-1]['coin']} {trades[-1]['entry_time'][:10]}"),
    ]

    cache = {}
    ROW_START = 3
    for pi, (trade, title) in enumerate(picks):
        r = ROW_START + pi * 22

        ws.cell(r, 1, title).font = Font(bold=True, size=12, color=TEXT_MAIN)
        ws.cell(r+1, 1, f"入场: {trade['entry_price']} | 离场: {trade['exit_price']} | "
                    f"PnL: {trade['pnl_usdt']:+}U | MAE: {trade.get('mae_pct',0)}% | MFE: {trade.get('mfe_pct',0)}%")
        ws.cell(r+1, 1).font = Font(size=9, color=TEXT_MUTED)

        # Fetch 15m price data
        sym = trade["coin"] + "-SWAP"
        if sym not in cache:
            cache[sym] = fetch_historical(sym, "15m", 30)
        candles = cache.get(sym, [])

        if not candles:
            ws.cell(r+2, 1, "数据不可用").font = Font(size=9, color=LOSE_RED)
            continue

        ets = trade["entry_ts"]
        exit_dt = datetime.strptime(trade["exit_time"], "%Y-%m-%d %H:%M").replace(tzinfo=timezone.utc)
        xts = int(exit_dt.timestamp() * 1000)

        ei = xi = None
        for j, c in enumerate(candles):
            if ei is None and int(c["ts"]) >= ets: ei = j
            if int(c["ts"]) >= xts: xi = j; break
        if ei is None: continue

        ep = candles[ei]["c"]
        # Extend 48 bars before and 96 bars after exit
        start_i = max(0, ei - 48)
        end_i = min(len(candles), (xi or ei+96) + 96)

        ws.cell(r+2, 1, "K线").font = Font(size=8, color=TEXT_MUTED)
        ws.cell(r+2, 2, "价格").font = Font(size=8, color=TEXT_MUTED)
        ws.cell(r+2, 3, "归一化%").font = Font(size=8, color=TEXT_MUTED)

        for j in range(start_i, end_i):
            idx = j - start_i
            cdl = candles[j]
            price = cdl["c"]
            norm = (price / ep - 1) * 100
            ws.cell(r+3+idx, 1, idx).font = Font(size=7, color=TEXT_MUTED)
            ws.cell(r+3+idx, 2, round(price, 8)).font = Font(size=7)
            ws.cell(r+3+idx, 3, round(norm, 2)).font = Font(size=7)

        data_end = r + 3 + (end_i - start_i) - 1

        chart = LineChart()
        chart.title = None
        chart.style = 2
        chart.width = 20
        chart.height = 10
        chart.legend = None

        dref = Reference(ws, min_col=3, min_row=r+2, max_col=3, max_row=data_end)
        cref = Reference(ws, min_col=1, min_row=r+3, max_row=data_end)
        chart.add_data(dref, titles_from_data=True)
        chart.set_categories(cref)

        # Color: green if profitable
        color = WIN_GREEN if trade["pnl_usdt"] > 0 else LOSE_RED
        chart.series[0].graphicalProperties.line.solidFill = color

        # Entry and exit markers
        ei_rel = ei - start_i
        ws.cell(r+3+ei_rel, 4, "←入场").font = Font(size=7, color="0d6efd", bold=True)
        if xi:
            xi_rel = xi - start_i
            ws.cell(r+3+xi_rel, 4, "←离场").font = Font(size=7, color=LOSE_RED, bold=True)

        ws.add_chart(chart, f"F{r}")

        # Hide price data
        for j in range(r+2, data_end+1):
            for c in range(1, 4):
                ws.cell(j, c).value = None

def _generate_html(trades, month_label, output_path):
    """生成网页版仪表盘"""
    import json as _json
    from collections import defaultdict as _dd

    data = _json.dumps(trades, ensure_ascii=False)
    total = len(trades)
    wins = sum(1 for t in trades if t["pnl_usdt"] > 0)
    total_pnl = sum(t["pnl_usdt"] for t in trades)
    win_rate = round(wins / total * 100) if total else 0

    coin_stats = _dd(lambda: {"pnl": 0, "count": 0})
    for t in trades:
        coin_stats[t["coin"]]["pnl"] += t["pnl_usdt"]
        coin_stats[t["coin"]]["count"] += 1

    daily_pnl = _dd(float)
    for t in trades:
        daily_pnl[t["entry_time"][:10]] += t["pnl_usdt"]
    daily_data = []
    cum = 0
    for day, pnl in sorted(daily_pnl.items()):
        cum += pnl
        daily_data.append({"day": day[5:], "pnl": round(pnl, 1), "cum": round(cum, 1)})

    mae_vals = [t.get("mae_pct", 0) for t in trades]
    mfe_vals = [t.get("mfe_pct", 0) for t in trades]
    avg_mae = sum(abs(v) for v in mae_vals) / len(mae_vals)
    avg_mfe = sum(mfe_vals) / len(mfe_vals)

    coins = [{"name": k, "pnl": round(v["pnl"], 1), "count": v["count"]} for k, v in coin_stats.items()]
    coins.sort(key=lambda x: x["pnl"], reverse=True)
    coins_json = _json.dumps(coins, ensure_ascii=False)
    daily_json = _json.dumps(daily_data, ensure_ascii=False)
    excel_name = f"交易记录_{month_label:02s}.xlsx" if len(str(month_label)) <= 2 else f"交易记录_{month_label}.xlsx"

    GREEN = "#198754"; RED = "#dc3545"; BLUE = "#0d6efd"

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>交易日记 · {month_label}月</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root{{--bg:#f8f9fa;--card:#fff;--text:#212529;--muted:#6c757d;--green:{GREEN};--red:{RED};--blue:{BLUE};--border:#dee2e6;--hover:#f1f3f5}}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','PingFang SC',sans-serif;background:var(--bg);color:var(--text);padding:16px;max-width:1200px;margin:0 auto}}
h1{{font-size:22px;margin-bottom:4px}}
.sub{{color:var(--muted);font-size:13px;margin-bottom:16px}}
.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:10px;margin-bottom:16px}}
.card{{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:14px;text-align:center}}
.card .label{{font-size:11px;color:var(--muted);margin-bottom:4px}}
.card .value{{font-size:24px;font-weight:700}}
.card .sub{{font-size:11px;color:var(--muted);margin-top:2px}}
.green{{color:var(--green)}}.red{{color:var(--red)}}.blue{{color:var(--blue)}}
.charts{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:16px}}
@media(max-width:768px){{.charts{{grid-template-columns:1fr}}}}
.chart-box{{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:14px}}
.chart-box h3{{font-size:13px;margin-bottom:8px;color:var(--muted)}}
.chart-wrap{{position:relative;height:240px;width:100%}}
canvas{{width:100%!important;height:100%!important}}
table{{width:100%;border-collapse:collapse;font-size:12px;background:var(--card);border-radius:8px;overflow:hidden;border:1px solid var(--border)}}
th{{background:#e9ecef;padding:8px 10px;text-align:left;font-weight:600;color:var(--muted);border-bottom:2px solid var(--blue);font-size:11px;white-space:nowrap}}
td{{padding:6px 10px;border-bottom:1px solid var(--border)}}
tr:hover td{{background:var(--hover)}}
tr:nth-child(even) td{{background:var(--bg)}}
tr:nth-child(even):hover td{{background:var(--hover)}}
.badge{{display:inline-block;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:600}}
.badge-win{{background:#d1e7dd;color:var(--green)}}.badge-lose{{background:#f8d7da;color:var(--red)}}
.badge-long{{background:#d1e7dd;color:var(--green)}}.badge-short{{background:#f8d7da;color:var(--red)}}
.btn{{display:inline-block;padding:8px 18px;border-radius:6px;font-size:13px;font-weight:600;text-decoration:none;margin-right:8px;cursor:pointer}}
.btn-down{{background:var(--green);color:#fff}}.btn-down:hover{{opacity:.9}}
.toolbar{{display:flex;justify-content:space-between;align-items:center;margin:16px 0 8px;flex-wrap:wrap;gap:8px}}
.toolbar input{{padding:6px 12px;border:1px solid var(--border);border-radius:6px;font-size:12px;outline:none;width:200px}}
.toolbar input:focus{{border-color:var(--blue)}}
.toolbar select{{padding:6px 12px;border:1px solid var(--border);border-radius:6px;font-size:12px;outline:none;background:var(--card)}}
</style>
</head>
<body>

<h1>📊 {month_label}月交易记录</h1>
<div class="sub">OKX 永续合约 · 2x杠杆逐仓 · 自动更新 · {total}笔</div>

<div class="cards">
    <div class="card"><div class="label">总交易</div><div class="value">{total}</div><div class="sub">笔</div></div>
    <div class="card"><div class="label">胜率</div><div class="value {"green" if win_rate>=50 else "red"}">{win_rate}%</div><div class="sub">{wins}赢/{total-wins}负</div></div>
    <div class="card"><div class="label">净P&L</div><div class="value {"green" if total_pnl>0 else "red"}">{total_pnl:+.1f}</div><div class="sub">USDT</div></div>
    <div class="card"><div class="label">平均MAE</div><div class="value red">{avg_mae:.1f}%</div><div class="sub">最大不利偏移</div></div>
    <div class="card"><div class="label">平均MFE</div><div class="value green">{avg_mfe:.1f}%</div><div class="sub">最大有利偏移</div></div>
</div>

<div style="margin-bottom:12px">
    <a href="{excel_name}" class="btn btn-down" download>⬇ 下载 Excel</a>
</div>

<div class="charts">
    <div class="chart-box"><h3>累计 P&L 走势</h3><div class="chart-wrap"><canvas id="chartCum"></canvas></div></div>
    <div class="chart-box"><h3>品种 P&L</h3><div class="chart-wrap"><canvas id="chartCoin"></canvas></div></div>
    <div class="chart-box"><h3>MAE vs MFE 散点</h3><div class="chart-wrap"><canvas id="chartMAE"></canvas></div></div>
    <div class="chart-box"><h3>日盈亏</h3><div class="chart-wrap"><canvas id="chartDay"></canvas></div></div>
</div>

<div class="toolbar">
    <input type="text" id="search" placeholder="搜索品种..." oninput="renderTable()">
    <select id="filterDir" onchange="renderTable()"><option value="all">全部方向</option><option value="做多">做多</option><option value="做空">做空</option></select>
    <select id="filterResult" onchange="renderTable()"><option value="all">全部结果</option><option value="win">盈利</option><option value="lose">亏损</option></select>
    <span style="font-size:11px;color:var(--muted)" id="rowCount"></span>
</div>

<table><thead><tr>
<th>日期</th><th>Coin</th><th>方向</th><th>1H</th><th>4H</th><th>1D</th><th>S1H</th><th>S4H</th><th>S1D</th>
<th>入场</th><th>离场</th><th>P&L%</th><th>P&L(U)</th><th>MAE</th><th>MFE</th><th>结果</th>
</tr></thead><tbody></tbody></table>

<script>
var D={data};
(function(){{
    var C=document.getElementById('chartCum').getContext('2d');
    var d={daily_json};
    new Chart(C,{{type:'line',data:{{labels:d.map(function(x){{return x.day}}),datasets:[{{label:'累计',data:d.map(function(x){{return x.cum}}),borderColor:'#0d6efd',backgroundColor:'rgba(13,110,253,.1)',fill:true,tension:.3,pointRadius:0}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{maxTicksLimit:15,color:'#6c757d'}}}},y:{{ticks:{{color:'#6c757d',callback:function(v){{return v+'U'}}}}}}}}}}}});

    var C2=document.getElementById('chartCoin').getContext('2d');
    var coins={coins_json};
    coins.sort(function(a,b){{return b.pnl-a.pnl}});
    new Chart(C2,{{type:'bar',data:{{labels:coins.map(function(c){{return c.name+' ('+c.count+')'}}),datasets:[{{label:'P&L USDT',data:coins.map(function(c){{return c.pnl}}),backgroundColor:coins.map(function(c){{return c.pnl>0?'rgba(25,135,84,.7)':'rgba(220,53,69,.7)'}}),borderColor:coins.map(function(c){{return c.pnl>0?'#198754':'#dc3545'}}),borderWidth:1,borderRadius:3}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{color:'#6c757d'}}}},y:{{ticks:{{color:'#6c757d',callback:function(v){{return v+'U'}}}}}}}}}}}});

    var C3=document.getElementById('chartMAE').getContext('2d');
    var m=D.map(function(t){{return{{x:t.mae_pct||0,y:t.mfe_pct||0,coin:t.coin}}}});
    new Chart(C3,{{type:'scatter',data:{{datasets:[{{label:'',data:m,backgroundColor:m.map(function(d){{return d.y>0?'rgba(25,135,84,.5)':'rgba(220,53,69,.5)'}}),pointRadius:5}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:function(c){{return c.raw.coin+' MAE:'+c.raw.x+'% MFE:'+c.raw.y+'%'}}}}}}}},scales:{{x:{{title:{{display:true,text:'MAE %',color:'#6c757d'}},ticks:{{color:'#6c757d',callback:function(v){{return v+'%'}}}}}},y:{{title:{{display:true,text:'MFE %',color:'#6c757d'}},ticks:{{color:'#6c757d',callback:function(v){{return v+'%'}}}}}}}}}}}});

    var C4=document.getElementById('chartDay').getContext('2d');
    new Chart(C4,{{type:'bar',data:{{labels:d.map(function(x){{return x.day}}),datasets:[{{label:'日P&L',data:d.map(function(x){{return x.pnl}}),backgroundColor:d.map(function(x){{return x.pnl>0?'rgba(25,135,84,.6)':'rgba(220,53,69,.6)'}}),borderWidth:1,borderRadius:3}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{maxTicksLimit:12,color:'#6c757d'}}}},y:{{ticks:{{color:'#6c757d',callback:function(v){{return v+'U'}}}}}}}}}}}});
}})();

function renderTable(){{
    var q=document.getElementById('search').value.toLowerCase();
    var d=document.getElementById('filterDir').value;
    var r=document.getElementById('filterResult').value;
    var f=D.filter(function(t){{
        if(q&&!t.coin.toLowerCase().includes(q))return false;
        if(d!=='all'&&t.direction!==d)return false;
        if(r==='win'&&t.pnl_usdt<=0)return false;
        if(r==='lose'&&t.pnl_usdt>=0)return false;
        return true;
    }});
    document.getElementById('rowCount').textContent=f.length+'/'+D.length+'笔';
    var h='';
    f.forEach(function(t){{
        var db=t.direction==='做多'?'badge-long':'badge-short';
        var rb=t.pnl_usdt>0?'badge-win':'badge-lose';
        h+='<tr><td>'+t.entry_time.substr(0,16)+'</td>'
          +'<td><strong>'+t.coin+'</strong></td>'
          +'<td><span class="badge '+db+'">'+t.direction+'</span></td>'
          +'<td class="'+(t.trend_1h==='多'?'green':'red')+'">'+t.trend_1h+'</td>'
          +'<td class="'+(t.trend_4h==='多'?'green':'red')+'">'+t.trend_4h+'</td>'
          +'<td class="'+(t.trend_1d==='多'?'green':'red')+'">'+t.trend_1d+'</td>'
          +'<td>'+t.srsi_1h+'</td><td>'+t.srsi_4h+'</td><td>'+t.srsi_1d+'</td>'
          +'<td>'+t.entry_price+'</td><td>'+t.exit_price+'</td>'
          +'<td class="'+(t.pnl_usdt>0?'green':'red')+'">'+t.pnl_pct.toFixed(1)+'%</td>'
          +'<td class="'+(t.pnl_usdt>0?'green':'red')+'">'+t.pnl_usdt.toFixed(2)+'</td>'
          +'<td class="red">'+(t.mae_pct||0).toFixed(1)+'%</td>'
          +'<td class="green">'+(t.mfe_pct||0).toFixed(1)+'%</td>'
          +'<td><span class="badge '+rb+'">'+(t.pnl_usdt>0?'盈利':'亏损')+'</span></td></tr>';
    }});
    document.querySelector('table tbody').innerHTML=h;
}}
renderTable();
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)


def main():
    now = datetime.now(timezone.utc) + timedelta(hours=8)  # CST
    month = int(sys.argv[1]) if len(sys.argv) > 1 else now.month
    year = now.year if month <= now.month else now.year

    print(f"🔄 拉取 {year}年{month}月交易记录...")
    try:
        start_ts, end_ts = get_month_range(year, month)
        new_trades = fetch_trades(start_ts, end_ts)
    except Exception as e:
        print(f"❌ OKX API 拉取失败: {e}")
        # Try to continue with existing data
        new_trades = []
    print(f"   OKX返回 {len(new_trades)} 笔已平仓交易")

    # 补充指标 (失败不影响主流程)
    if new_trades:
        try:
            new_trades = enrich_with_indicators(new_trades)
        except Exception as e:
            print(f"⚠️ 指标计算失败: {e}，将跳过指标列")

    # 合并已有
    existing = load_existing(DATA_FILE)
    merged, added = merge_trades(existing, new_trades)

    # 保存
    with open(DATA_FILE, "w") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)

    # 生成 Excel
    excel_path = EXCEL_FILE.format(month=f"{month:02d}")
    generate_excel(merged, excel_path, f"{month}")

    print(f"\n✅ 总计 {len(merged)} 笔交易")
    print(f"   本次新增 {len(added)} 笔")
    if added:
        for t in added:
            print(f"     + {t['entry_time'][:10]} {t['coin']} {t['direction']} PnL:{t['pnl_usdt']:+}U")
    print(f"   Excel: {excel_path}")

    # 返回路径
    result_file = os.path.join(OUTPUT_DIR, ".last_run.json")
    with open(result_file, "w") as f:
        json.dump({"excel": excel_path, "total": len(merged), "added": len(added)}, f)


if __name__ == "__main__":
    main()
